# -*- coding: utf-8-*-
import os
import re
import sys
import openpyxl as op
import xlrd, xlwt

SELF_NAME = os.path.basename(sys.argv[0]).split(".")[0]


class MainSearch(object):
    """
            searching specific files
            author: guoyuqiang
            :param path:root path
            :param name:filename searched
            :param only:stop searching after find 1 file
            :param mdir:file or dir,true is file
            :param mode:searching order > 0 or 1,0 is hor,1 is ver
    """
    _min = _max = -1

    def __init__(self, path=None, name='[\S\s]*', only=False, mdir=False, mode=1):

        self.mode = mode
        self.only = only
        self.name = name
        self.path = path
        self.mdir = mdir
        self._result_files = []

    def _search_files(self):

        _paths = [os.path.join(self.path, x) for x in os.listdir(self.path)]
        while len(_paths) != 0:
            _f = _paths.pop(0)
            if self.mdir:
                if os.path.isdir(_f):
                    self._result_files.append(_f)
                    for p in os.listdir(_f):
                        p = os.path.join(_f, p)
                        if os.path.isdir(p):
                            _paths.append(p)
                            self._result_files.append(p)
            else:
                if os.path.isfile(_f):
                    size = os.path.getsize(_f)
                    if self._min != -1 and self._max != -1:
                        if size < self._min or size > self._max:
                            continue
                    elif self._min != -1:
                        if size < self._min:
                            continue
                    elif self._max != -1:
                        if size > self._max:
                            continue
                    is_match = re.match(self.name, os.path.split(_f)[1])
                    if is_match:
                        self._result_files.append(_f)
                        if self.only:
                            break
                elif os.path.isdir(_f):
                    if self.mode == 0:
                        for p in os.listdir(_f):
                            _paths.append(os.path.join(_f, p))
                    else:
                        self.path = _f
                        self._search_files()

    def set_min(self, i):
        self._min = i
        return self

    def set_name(self, name):
        self.name = name

    def set_max(self, i):
        self._max = i
        return self

    def set_path(self, path):
        self.path = path

    def start(self):
        self._result_files.clear()
        self._search_files()

    def get_files(self):
        return self._result_files

    def removes(self):
        for _f in self._result_files:
            os.remove(_f)

    def together(self, path):
        """
        :param path: put them all in here
        :return:
        """
        for f in self._result_files:
            if not os.path.exists(path):
                os.makedirs(path)
            os.rename(f, path + os.path.split(f)[1])

    def rename(self, before, after):
        for fi in self._result_files:
            p = fi[:fi.rfind(os.sep) + 1]
            n = os.path.split(fi)[1]
            is_fit = re.search(before, n)
            if is_fit:
                before = is_fit.group()
                if before == after:
                    continue
                # os.chdir(os.path.split(fi)[0])
                try:
                    os.rename(fi, p + n.replace(before, after))
                except WindowsError:
                    print(fi)


def save_memory(filename, content):
    mp = filename[:filename.rfind(os.sep)]
    if not os.path.exists(mp):
        os.makedirs(mp)
    with open(filename, "a+") as f:
        f.write(content + '\n')


# def add_title(_paths):
#     for p in _paths:
#         oldWb = xlrd.open_workbook(p)
#         oldWbS = oldWb.sheet_by_index(0)
#         newWb = copy(oldWb)
#         newWs = newWb.get_sheet(0)
#         inserColNo = 0
#         newWs.write(0, inserColNo, "项目名称")
#         for i in range(1, oldWbS.nrows):
#             newWs.write(i, inserColNo, os.path.basename(p)[:-5])
#
#         for rowIndex in range(inserColNo, oldWbS.nrows):
#             for colIndex in range(oldWbS.ncols):
#                 newWs.write(rowIndex, colIndex + 1, oldWbS.cell(rowIndex, colIndex).value)
#         newWb.save(p.replace('xlsx', 'xls'))
# newWb.save(p)

def rename2z(path):
    for p in path:
        start = p[:p.rfind('_') + 1]
        end = p[p.rfind('.'):]
        fnl = start + '0' + end
        print(p)
        print(fnl)
        # print('')
        # path = p[:p.rfind('/') + 1] + p[p.rfind('_') + 1:p.rfind('.')] + p[p.rfind('/'):]
        # print(path)
        # if not os.path.exists(path[:path.rfind('/')]):
        #     os.mkdir(path[:path.rfind('/')])
        # os.rename(p, fnl)


# 筛选手机号
def do_xl(paths):
    if not os.path.exists('result'):
        os.mkdir('result')
    for path in paths:
        print(path)
        file = op.Workbook(True)
        wb = xlrd.open_workbook(path)
        for name in wb.sheets():
            print(name)
            write = file.create_sheet(name.name)
            ws = wb.sheet_by_name(name.name)
            # 获取title
            t1 = t2 = -1
            m_row = [x.value for x in ws.row(0)]
            write.append(m_row)
            for i, e in enumerate(ws.row(0)):
                if e.value == '事件标签':
                    t1 = i
                elif e.value == '项目名称':
                    t2 = i

            r = 1
            for rows in ws.get_rows():
                if check_no1(rows[t1].value) and check_no2(rows[t2].value):
                    # for i, row in enumerate(rows):
                    m_row = [x.value for x in rows]
                    write.append(m_row)
                else:
                    r -= 1
                r += 1
        file.save('.%sresult%s' % (os.sep, path[path.rfind(os.sep):]))


def check_no1(num):
    if len(num) == 7 or len(num) == 11:
        try:
            int(num)
            return True
        except ValueError:
            return False
    else:
        return False


def check_no2(num):
    if num.strip() == '':
        return False
    else:
        return True


# 生成报表
def scbb(paths):
    if not os.path.exists('report'):
        os.mkdir('report')
    for path in paths:
        data = {}
        print(path)
        file = op.Workbook(True)
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        for rows in ws.get_rows():
            if rows[0].value == '月份':
                continue
            tmp = [rows[1].value, rows[6].value, rows[2].value, rows[7].value, 1]
            if tmp[3].strip() == '':
                continue
            if rows[0].value not in data.keys():
                data[rows[0].value] = [tmp]
                continue
            is_find = False
            for i, item in enumerate(data[rows[0].value]):
                if tmp[:-1] == item[:-1]:
                    data[rows[0].value][i][-1] += 1
                    is_find = True
                    break
            if not is_find:
                data[rows[0].value].append(tmp)
        for key, value in data.items():
            wt = file.create_sheet(key)
            wt.append(['项目', '项目名称', '媒体', '媒体名称', '汇总'])
            for v in value:
                wt.append(v)
        file.save('report%s' % path[path.rfind(os.sep):])


if __name__ == '__main__':
    _path = sys.argv[1]
    main = MainSearch(_path, u'[\S\s]*', mdir=True)
    main.start()
    for o in main.get_files():
        print(o)
    # scbb(main.get_files())
    # do_xl(main.get_files())
    # rename2z(main.get_files())
    # print(copy_list([1, 2, 'adsf',3, 4]))
